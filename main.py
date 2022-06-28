import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import time
import openpyxl
import pandas as pd
import sys
import datetime
import config


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(len(sys.argv))
        print("Usage : python main.py [excel file]")
        sys.exit(0)
    args = sys.argv[1]

file_read = str(args)
file_index = "index.txt"
config = config.Config()


try:
    f = open(file_read, 'r')
except:
    print("[ERROR] File Does not exist")
    sys.exit(0)

try:
    f2 = open(file_index, 'wb')
except:
    print("[ERROR] File Could not be created exist")
    sys.exit(0)


start_index = 1;
with open(file_index, "r") as myfile:
    index = myfile.readline()
    if index != "":
        start_index = int(index)


# Create a new log file with current date:time
file_write = str(datetime.datetime.now().strftime("%m-%d-%y-%H-%M-%S")) + ".xlsx"

# the workbook for Excel file to read from
wrkbk = openpyxl.load_workbook(file_read)

# the workbook file to write to
wrkb2 = openpyxl.Workbook()
wrkb2.save(file_write)
sh = wrkbk.active
sh2 = wrkb2.active
sh2.cell(row=1, column=1).value = "Names"
sh2.cell(row=1, column=2).value = "Emails"
sh2.cell(row=1, column=3).value = "Message"
wrkb2.save(file_write)



# configuration file
server = config.get_server()
port = int(config.get_port())
sender = config.get_sender()
password = config.get_password()
milestone = config.get_milestone()
delimeter = config.get_delimeter()
max_rate = int(config.get_max_submit_rate())

msg = MIMEMultipart('alternative')
msg['Subject'] = config.get_subject()
msg['From'] = config.get_sender()
attachement = config.get_attached_file()

username_column=1
email_column = 2

if attachement:
    with open(attachement, "rb") as f:
        attach = MIMEApplication(f.read(), _subtype="pdf")
    attach.add_header('Content-Disposition', 'attachment', filename=str(attachement))
    msg.attach(attach)



html_const = config.read_message()
file_write_index = 1

html_body = MIMEText(html_const, 'html', _charset="UTF-8")
msg.attach(html_body)


try:
    smtpObj = smtplib.SMTP(server, port)
    smtpObj.ehlo()
    smtpObj.starttls()
    smtpObj.login(sender, password)
    print("Login successful..")

    # smtpObj.set_debuglevel(1)


    for i in range(start_index, sh.max_row + 1):



        username = str(sh.cell(row=i + 1, column=username_column).value)
        email = str(sh.cell(row=i + 1, column=email_column).value)

        # should read html from file


        try:

            # First of All in loop
            if i % max_rate == 0:
                time.sleep(60)

            # the problem is here
            #check that the email is legit
            res = smtpObj.sendmail(sender, email, msg.as_string())
            if(res == {}):
                print("Message successfuly sent to : " + username + ", " + email)

            sh2.cell(row=file_write_index + 1, column=1).value = username
            sh2.cell(row=file_write_index + 1, column=2).value = email
            sh2.cell(row=file_write_index + 1, column=3).value = html_const
            print("Index : " + str(file_write_index) + ", Name : " + username + ", Email : " + email)
            file_write_index += 1

            with open(file_index, "w") as myfile:
                myfile.write(str(i))
            #f2.write(str(i).encode())
            #f2.truncate()
            wrkb2.save(file_write)

        # If program fails
        except smtplib.SMTPException:

            # save file
            wrkb2.save(file_write)
            f2.write(str(i).encode())

            i += 1

            # reset connection
            smtpObj.close()
            # wait for max submittion rate's time
            time.sleep(60)
            smtpObj = smtplib.SMTP('192.168.12.61', 587)
            smtpObj.ehlo()
            smtpObj.starttls()
            smtpObj.login(sender, password)
            continue


except smtplib.SMTPException:
    print("Error: unable to send email")

    smtpObj.close()
wrkb2.save(file_write)
smtpObj.close()
f.close()
f2.close()

